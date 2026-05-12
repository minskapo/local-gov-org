"""기후·환경 관련 행정기구 현황 → 엑셀

포함: 사업소, 본청 기구 등 일반 행정기구
제외: 연구원, 연구소, 연구실 (비일반 행정기관)
기초명 형식: "서울특별시 마포구" → "마포구" (마지막 단어)
"""
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
OUT_DIR = BASE_DIR / "data" / "processed" / "by_region"
EXCEL_PATH = BASE_DIR / "data" / "processed" / "기후환경_행정기구_현황.xlsx"

CLIMATE_KW = frozenset(["기후", "환경", "생태", "탄소"])

# 연구원·연구소·연구실 등 일반 행정기관이 아닌 기구 제외
EXCLUDE_KW = frozenset(["연구원", "연구소", "연구실"])


def is_climate(name: str) -> bool:
    return any(kw in name for kw in CLIMATE_KW)


def is_excluded(name: str) -> bool:
    return any(kw in name for kw in EXCLUDE_KW)


def short_gicheo(full_name: str) -> str:
    """'서울특별시 마포구' → '마포구'"""
    parts = full_name.strip().split()
    return parts[-1] if len(parts) > 1 else full_name


def make_comment(unit: dict):
    """분장사무_항목 또는 분장사무_원문으로 Excel 메모 생성. 내용 없으면 None."""
    items = unit.get("분장사무_항목") or []
    raw   = (unit.get("분장사무_원문") or "").strip()

    if items:
        text = "\n".join(f"{i+1}. {item}" for i, item in enumerate(items))
    elif raw:
        text = re.sub(r"^제\d+조(?:의\d+)?\([^)]{0,120}\)", "", raw).strip()
    else:
        return None

    if len(text) > 3000:
        text = text[:2997] + "…"

    c = Comment(text, "소관사무")
    c.width  = 320
    c.height = min(40 + len(items or text.split("\n")) * 18, 400)
    return c


def load_all():
    records = []
    for f in sorted(OUT_DIR.glob("*.json")):
        with open(f, encoding="utf-8") as fh:
            records.append(json.load(fh))
    return records


def get_climate_rows(d):
    """기후환경 관련 상위기구 행 목록 반환.
    각 행: {parent, children, climate_parent(bool)}
    - climate_parent=True: 상위기구명 자체가 기후환경 관련
    - climate_parent=False: 상위기구는 비기후이지만 하위기구 중 기후환경 기구 존재
    연구원·연구소·연구실은 제외.
    """
    rows = []
    for unit in d.get("structure", []):
        name = unit.get("name", "")
        if is_excluded(name):
            continue
        child_objs = unit.get("children", [])
        children   = [c.get("name", "") for c in child_objs]
        if is_climate(name):
            rows.append({"parent": name, "parent_obj": unit,
                         "children": children, "child_objs": child_objs,
                         "climate_parent": True})
        elif any(is_climate(c) for c in children):
            rows.append({"parent": name, "parent_obj": unit,
                         "children": children, "child_objs": child_objs,
                         "climate_parent": False})
    return rows


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_cell(ws, row, col, value, bg, bold=False, align="center", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.border = thin_border()
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return cell


def main():
    print("데이터 로딩 중...")
    records = load_all()
    print(f"  {len(records)}개 지자체 로드 완료")

    # 코드 → 광역명 매핑
    code_to_name = {d["region"]["code"]: d["region"]["name"] for d in records}

    # 최대 하위기구 수 파악
    max_children = max(
        (len(row["children"]) for d in records for row in get_climate_rows(d)),
        default=0,
    )
    print(f"  최대 하위기구 수: {max_children}")

    wb = Workbook()
    ws = wb.active
    ws.title = "기후환경 행정기구 현황"

    # ── 헤더 ─────────────────────────────────────────────
    FIXED = [
        ("지자체코드",        9),
        ("광역명",           14),
        ("기초명",           14),
        ("기후환경\n관련여부", 9),
        ("상위기구",         22),
    ]
    HDR_BG = "1E3A5F"
    for col, (title, width) in enumerate(FIXED, 1):
        apply_cell(ws, 1, col, title, HDR_BG, bold=True, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = width
        ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF", size=10)

    START = len(FIXED) + 1
    for i in range(1, max_children + 1):
        col = START + i - 1
        apply_cell(ws, 1, col, f"하위기구{i}", HDR_BG, bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 18
        ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF", size=10)
    ws.row_dimensions[1].height = 32

    # ── 데이터 ───────────────────────────────────────────
    # 색상
    BG_PARENT_CLIMATE  = "DCFCE7"  # 연초록: 상위기구 자체가 기후환경
    BG_PARENT_INDIRECT = "FEF9C3"  # 연노랑: 하위기구에만 기후환경 있음
    BG_CHILD_CLIMATE   = "BFDBFE"  # 연파랑: 기후환경 관련 하위기구
    BG_CHILD_OTHER     = "F8FAFC"  # 연회색: 비기후 하위기구

    # 2개 이상 행이 나오는 지자체 코드 사전 파악
    multi_codes = {
        d["region"]["code"]
        for d in records
        if len(get_climate_rows(d)) > 1
    }

    data_row = 2
    total_rows = 0

    for d in records:
        reg = d["region"]
        code = reg.get("code", "")
        name = reg.get("name", "")
        level = reg.get("level", "")
        parent_code = reg.get("parent") or ""

        if level == "광역":
            gwangyeok = name
            gicheo = ""
        else:
            gwangyeok = code_to_name.get(parent_code, "")
            if not gwangyeok:
                parts = name.split()
                gwangyeok = parts[0] if len(parts) > 1 else name
            gicheo = short_gicheo(name)  # "서울특별시 마포구" → "마포구"

        climate_rows = get_climate_rows(d)
        if not climate_rows:
            BG_EMPTY = "F1F5F9"
            fixed_vals = [code, gwangyeok, gicheo, "", ""]
            aligns     = ["center", "left", "left", "center", "left"]
            for col, (val, al) in enumerate(zip(fixed_vals, aligns), 1):
                apply_cell(ws, data_row, col, val, BG_EMPTY, align=al)
            data_row += 1
            total_rows += 1
            continue

        is_multi = code in multi_codes
        for cr in climate_rows:
            if is_multi:
                row_bg      = "FDE68A"  # 연황색: 중복 지자체 행 전체
                child_bg_kw = "FDE68A"
                child_bg_ot = "FDE68A"
            else:
                row_bg      = BG_PARENT_CLIMATE if cr["climate_parent"] else BG_PARENT_INDIRECT
                child_bg_kw = BG_CHILD_CLIMATE
                child_bg_ot = BG_CHILD_OTHER
            label = "○" if cr["climate_parent"] else "△"

            fixed_vals = [code, gwangyeok, gicheo, label, cr["parent"]]
            aligns     = ["center", "left", "left", "center", "left"]
            for col, (val, al) in enumerate(zip(fixed_vals, aligns), 1):
                cell = apply_cell(ws, data_row, col, val, row_bg, align=al)
                if col == 5:
                    cmt = make_comment(cr["parent_obj"])
                    if cmt:
                        cell.comment = cmt

            for i, (child, child_obj) in enumerate(zip(cr["children"], cr["child_objs"])):
                col = START + i
                child_bg = child_bg_kw if is_climate(child) else child_bg_ot
                cell = apply_cell(ws, data_row, col, child, child_bg, align="left")
                cmt = make_comment(child_obj)
                if cmt:
                    cell.comment = cmt

            data_row += 1
            total_rows += 1

    ws.freeze_panes = "A2"
    total_cols = START + max_children - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}1"

    print(f"저장 중: {EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    size_kb = EXCEL_PATH.stat().st_size / 1024
    print(f"완료! {total_rows}행 작성")
    print(f"파일 크기: {size_kb:.0f} KB")
    print(f"\n[참고] 사용 키워드: {', '.join(sorted(CLIMATE_KW))}")


if __name__ == "__main__":
    main()
