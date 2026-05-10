"""243개 지자체 행정기구 데이터 → 단일 엑셀 파일 생성"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
OUT_DIR = BASE_DIR / "data" / "processed" / "by_region"
EXCEL_PATH = BASE_DIR / "data" / "processed" / "지자체_행정기구_전체.xlsx"

# ── 색상 팔레트 ───────────────────────────────────────────
C = {
    "header_bg":   "2563EB",  # 파란색 헤더
    "header_fg":   "FFFFFF",
    "gwangyeok":   "DBEAFE",  # 광역 행 배경
    "gicheo":      "F0FDF4",  # 기초 행 배경
    "unit_lv1":    "EFF6FF",  # 레벨1 기구
    "unit_lv2":    "FFFFFF",  # 레벨2 기구
    "type_국":     "DBEAFE",
    "type_실":     "F3E8FF",
    "type_본부":   "FEF9C3",
    "type_단":     "FFEDD5",
    "type_관":     "DCFCE7",
    "type_직속":   "F1F5F9",
    "type_담당관": "FDF4FF",
    "type_과":     "FFFFFF",
    "border":      "CBD5E1",
    "alt_row":     "F8FAFC",
}

TYPE_COLORS = {
    "국": "DBEAFE", "실": "F3E8FF", "본부": "FEF9C3",
    "단": "FFEDD5", "관": "DCFCE7", "직속기관": "F1F5F9",
    "담당관": "FDF4FF", "과": "FFFFFF", "처": "FEE2E2",
}


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def border_thin():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def header_font():
    return Font(bold=True, color=C["header_fg"], size=10)


def load_all():
    files = sorted(OUT_DIR.glob("*.json"))
    records = []
    for f in files:
        with open(f, encoding="utf-8") as fh:
            d = json.load(fh)
        records.append(d)
    return records


def resolve_parent_name(reg: dict, code_to_name: dict) -> str:
    """parent 코드 → 광역명. 코드 체계 불일치 시 지자체명에서 추출."""
    parent_code = reg.get("parent") or ""
    if not parent_code:
        return ""
    name = code_to_name.get(parent_code, "")
    if not name:
        full_name = reg.get("name", "")
        parts = full_name.split()
        name = parts[0] if len(parts) > 1 else ""
    return name


# ── 시트 1: 지자체 요약 ──────────────────────────────────
def build_summary_sheet(ws, records, code_to_name=None):
    ws.title = "지자체 요약"

    COLS = [
        ("지자체코드", 10), ("지자체명", 18), ("광역/기초", 9),
        ("지자체유형", 12), ("상위지자체", 14), ("총정원(명)", 11),
        ("본청 최상위기구", 10), ("하위기구 합계", 10), ("직속기관 수", 10),
        ("국·실·본부", 10), ("과·담당관 (하위)", 13),
    ]

    # 헤더
    for col, (title, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill(C["header_bg"])
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 32

    row = 2
    for d in records:
        reg = d["region"]
        totals = d.get("totals", {})
        structure = d.get("structure", [])

        top_units = [n for n in structure if n.get("type") != "직속기관"]
        agencies = [n for n in structure if n.get("type") == "직속기관"]
        children_total = sum(len(n.get("children", [])) for n in top_units)
        주요기구 = [n for n in top_units if n.get("type") in ("국","실","본부","단","관","처")]
        하위기구 = sum(len(n.get("children",[])) for n in 주요기구)

        vals = [
            reg.get("code",""),
            reg.get("name",""),
            reg.get("level",""),
            reg.get("type",""),
            resolve_parent_name(reg, code_to_name or {}),
            totals.get("정원_총", 0) or 0,
            len(top_units),
            children_total,
            len(agencies),
            len(주요기구),
            하위기구,
        ]

        is_gwangyeok = reg.get("level") == "광역"
        row_bg = C["gwangyeok"] if is_gwangyeok else (
            C["alt_row"] if row % 2 == 0 else C["gicheo"]
        )

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill(row_bg)
            cell.border = border_thin()
            cell.alignment = Alignment(
                horizontal="center" if col not in (2, 5) else "left",
                vertical="center"
            )
            if col == 6 and isinstance(val, int):
                cell.number_format = "#,##0"
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


# ── 시트 2: 행정기구 전체 목록 ──────────────────────────
def build_units_sheet(ws, records, code_to_name=None):
    ws.title = "행정기구 목록"

    COLS = [
        ("지자체코드", 9), ("지자체명", 16), ("광역/기초", 8),
        ("지자체유형", 10), ("상위지자체", 12),
        ("계층", 5), ("상위기구", 16), ("기구유형", 9), ("기구명", 22),
        ("장직급", 10), ("하위기구수", 8), ("총정원", 8),
        ("키워드", 30),
    ]

    for col, (title, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill(C["header_bg"])
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28

    row = 2
    for d in records:
        reg = d["region"]
        totals = d.get("totals", {})
        structure = d.get("structure", [])

        code = reg.get("code","")
        name = reg.get("name","")
        level = reg.get("level","")
        rtype = reg.get("type","")
        parent = resolve_parent_name(reg, code_to_name or {})
        총정원 = totals.get("정원_총", 0) or 0

        for unit in structure:
            u_type = unit.get("type","")
            u_name = unit.get("name","")
            children = unit.get("children", [])
            grade = unit.get("head_grade","")
            keywords = ", ".join(unit.get("키워드_태그", []))
            bg = TYPE_COLORS.get(u_type, "FFFFFF")

            vals = [code, name, level, rtype, parent,
                    1, "", u_type, u_name,
                    grade, len(children), 총정원, keywords]

            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill(bg)
                cell.border = border_thin()
                cell.alignment = Alignment(
                    horizontal="center" if col not in (2,5,9,13) else "left",
                    vertical="center",
                    wrap_text=(col == 13),
                )
                if col == 12 and isinstance(val, int) and val > 0:
                    cell.number_format = "#,##0"
            # Bold for 레벨1 중요 기구
            if u_type in ("국","실","본부","단"):
                ws.cell(row=row, column=9).font = Font(bold=True, size=10)
            row += 1

            # 하위 기구 rows
            for child in children:
                c_type = child.get("type","")
                c_name = child.get("name","")
                c_grade = child.get("head_grade","")
                c_keywords = ", ".join(child.get("키워드_태그", []))
                c_bg = TYPE_COLORS.get(c_type, "FFFFFF")

                c_vals = [code, name, level, rtype, parent,
                          2, u_name, c_type, "  └ " + c_name,
                          c_grade, "", "", c_keywords]

                for col, val in enumerate(c_vals, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.fill = fill(c_bg)
                    cell.border = border_thin()
                    cell.alignment = Alignment(
                        horizontal="center" if col not in (2,5,7,9,13) else "left",
                        vertical="center",
                        wrap_text=(col == 13),
                    )
                row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


# ── 시트 3: 지자체별 통계 피벗 ──────────────────────────
def build_stats_sheet(ws, records):
    ws.title = "기구유형별 통계"

    # 유형별 집계
    TYPES = ["국", "실", "본부", "단", "관", "처", "담당관", "과", "직속기관"]
    COLS = ["지자체코드","지자체명","광역/기초","지자체유형","총정원"] + TYPES + ["전체기구수"]

    widths = [9,16,8,10,10] + [7]*len(TYPES) + [9]
    for col, (title, w) in enumerate(zip(COLS, widths), 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill(C["header_bg"])
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    row = 2
    for d in records:
        reg = d["region"]
        totals = d.get("totals", {})
        structure = d.get("structure", [])

        # 유형별 카운트 (레벨1 + 레벨2 모두 포함)
        type_count = {t: 0 for t in TYPES}
        total_units = 0
        for unit in structure:
            ut = unit.get("type","")
            if ut in type_count:
                type_count[ut] += 1
            total_units += 1
            for child in unit.get("children",[]):
                ct = child.get("type","")
                if ct in type_count:
                    type_count[ct] += 1
                total_units += 1

        is_gw = reg.get("level") == "광역"
        row_bg = C["gwangyeok"] if is_gw else (
            C["alt_row"] if row % 2 == 0 else "FFFFFF"
        )

        vals = [
            reg.get("code",""), reg.get("name",""),
            reg.get("level",""), reg.get("type",""),
            totals.get("정원_총", 0) or 0,
        ] + [type_count[t] for t in TYPES] + [total_units]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill(row_bg)
            cell.border = border_thin()
            cell.alignment = Alignment(
                horizontal="center" if col != 2 else "left",
                vertical="center"
            )
            if col == 5 and isinstance(val, int):
                cell.number_format = "#,##0"
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def main():
    print("데이터 로딩 중...")
    records = load_all()
    print(f"  {len(records)}개 지자체 로드 완료")

    code_to_name = {d["region"]["code"]: d["region"]["name"] for d in records}

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    print("시트 1: 지자체 요약 생성 중...")
    ws1 = wb.create_sheet()
    build_summary_sheet(ws1, records, code_to_name)

    print("시트 2: 행정기구 목록 생성 중...")
    ws2 = wb.create_sheet()
    build_units_sheet(ws2, records, code_to_name)

    print("시트 3: 기구유형별 통계 생성 중...")
    ws3 = wb.create_sheet()
    build_stats_sheet(ws3, records)

    print(f"저장 중: {EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    print(f"완료! 파일 크기: {EXCEL_PATH.stat().st_size / 1024:.0f} KB")


if __name__ == "__main__":
    main()
