"""지자체별 최상위 행정기구 1개 = 1행 엑셀 생성"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
OUT_DIR = BASE_DIR / "data" / "processed" / "by_region"
EXCEL_PATH = BASE_DIR / "data" / "processed" / "행정기구_상하위_목록.xlsx"

TYPE_BG = {
    "국": "DBEAFE", "실": "F3E8FF", "본부": "FEF9C3",
    "단": "FFEDD5", "관": "DCFCE7", "처": "FEE2E2",
    "직속기관": "F1F5F9", "담당관": "FDF4FF", "과": "FFFFFF",
}
CHILD_BG = "F8FAFC"
HDR_BG   = "1E3A5F"


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def cell_write(ws, row, col, value, bg, bold=False, align="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.border = thin_border()
    c.font = Font(bold=bold, color="FFFFFF" if bg == HDR_BG else "000000", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c


def load_all():
    records = []
    for f in sorted(OUT_DIR.glob("*.json")):
        with open(f, encoding="utf-8") as fh:
            records.append(json.load(fh))
    return records


def make_comment(unit: dict):
    """분장사무_항목 또는 분장사무_원문으로 Excel 메모 생성. 내용 없으면 None."""
    items = unit.get("분장사무_항목") or []
    raw   = (unit.get("분장사무_원문") or "").strip()

    if items:
        lines = [f"{i+1}. {item}" for i, item in enumerate(items)]
        text = "\n".join(lines)
    elif raw:
        # 헤더(제X조(...)) 제거 후 사용
        import re
        text = re.sub(r"^제\d+조(?:의\d+)?\([^)]{0,120}\)", "", raw).strip()
    else:
        return None

    if len(text) > 3000:
        text = text[:2997] + "…"

    c = Comment(text, "소관사무")
    c.width  = 320
    c.height = min(40 + len(items or text.split("\n")) * 18, 400)
    return c


def resolve_gwangyeok(reg, code_to_name):
    level = reg.get("level", "")
    name  = reg.get("name", "")
    if level == "광역":
        return name, ""
    parent_code = reg.get("parent") or ""
    gwangyeok = code_to_name.get(parent_code, "")
    if not gwangyeok:
        parts = name.split()
        gwangyeok = parts[0] if len(parts) > 1 else name
    return gwangyeok, name


def main():
    print("데이터 로딩 중...")
    records = load_all()
    print(f"  {len(records)}개 지자체 로드 완료")

    code_to_name = {d["region"]["code"]: d["region"]["name"] for d in records}

    # 최대 하위기구 수 파악
    max_children = max(
        (len(u.get("children", []))
         for d in records
         for u in d.get("structure", [])),
        default=0,
    )
    print(f"  최대 하위기구 수: {max_children}")

    wb = Workbook()
    ws = wb.active
    ws.title = "행정기구 상하위 목록"

    # ── 헤더 ─────────────────────────────────────────────────
    FIXED = [
        ("코드",   9),
        ("광역",  16),
        ("기초",  16),
        ("기구유형", 9),
        ("상위 행정기구명", 22),
    ]
    for col, (title, width) in enumerate(FIXED, 1):
        cell_write(ws, 1, col, title, HDR_BG, bold=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    START = len(FIXED) + 1
    for i in range(1, max_children + 1):
        col = START + i - 1
        cell_write(ws, 1, col, f"하위기구{i}", HDR_BG, bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.row_dimensions[1].height = 28

    # ── 데이터 ───────────────────────────────────────────────
    data_row = 2
    total_rows = 0

    for d in records:
        reg = d["region"]
        code = reg.get("code", "")
        gwangyeok, gicheo = resolve_gwangyeok(reg, code_to_name)

        for unit in d.get("structure", []):
            u_type    = unit.get("type", "")
            u_name    = unit.get("name", "")
            child_objs = unit.get("children", [])
            row_bg    = TYPE_BG.get(u_type, "FFFFFF")

            fixed_vals   = [code, gwangyeok, gicheo, u_type, u_name]
            fixed_aligns = ["center", "left", "left", "center", "left"]
            for col, (val, al) in enumerate(zip(fixed_vals, fixed_aligns), 1):
                bold = (col == 5 and u_type in ("국", "실", "본부", "단"))
                cell = cell_write(ws, data_row, col, val, row_bg, bold=bold, align=al)
                if col == 5:
                    cmt = make_comment(unit)
                    if cmt:
                        cell.comment = cmt

            for i, child in enumerate(child_objs):
                col  = START + i
                cell = cell_write(ws, data_row, col, child.get("name", ""), CHILD_BG, align="left")
                cmt  = make_comment(child)
                if cmt:
                    cell.comment = cmt

            data_row += 1
            total_rows += 1

    ws.freeze_panes = "A2"
    total_cols = START + max_children - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}1"

    print(f"저장 중: {EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    size_kb = EXCEL_PATH.stat().st_size / 1024
    print(f"완료! {total_rows}행, {size_kb:.0f} KB")


if __name__ == "__main__":
    main()
